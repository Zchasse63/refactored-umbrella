#!/usr/bin/env python3
"""Partner Excel backup — a product/competitor comparison workbook with live deal math.

Products & Competitors sheet: each product's economics (cost → target sell → FBA +
selling fees → net/unit → margin) is pinned in FROZEN left columns merged once per
product; Net and Margin are LIVE FORMULAS (edit Target Sell and they recompute). Its
competitors read to the right; lowest competitor price per group is shaded green.

Run: python3 scripts/build-backup.py   →   ~/Desktop/Portal-Backup-<date>.xlsx
"""
import json, os, urllib.request, datetime, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
env = {}
for line in open(os.path.join(ROOT, ".env.local")):
    if "=" in line and not line.strip().startswith("#"):
        k, v = line.split("=", 1); env[k.strip()] = v.strip().strip('"').strip("'")
URL, KEY = env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"]
def get(path):
    r = urllib.request.Request(URL + path, headers={"apikey": KEY, "Authorization": f"Bearer {KEY}"})
    return json.loads(urllib.request.urlopen(r, timeout=90).read())

SITE = "https://the-portal-sourcing.netlify.app"
OPEX = 0.34  # referral 15% + ads 15% + returns 4% (FBA is charged separately, in $)
prods = get("/rest/v1/products?select=id,external_ref,line,group_name,name,model,specs,our_cost&order=line,name")
comps = get("/rest/v1/competitors?status=eq.approved&select=product_id,title,retail_url,price,bsr,est_monthly_sales,review_count,fba_pick_pack_fee&order=est_monthly_sales.desc")
costs = json.load(open("/tmp/fs-costs.json")) if os.path.exists("/tmp/fs-costs.json") else {}
cby = {}
for c in comps: cby.setdefault(c["product_id"], []).append(c)

INK, SLATE, MUT, LINK = "FF1E293B", "FF475569", "FF64748B", "FF2563EB"
HUE = {"foodservice": "FF0E7490", "appliance": "FF6D28D9", "beauty": "FFBE185D"}
BAND = "FFEDF2F8"
LOWFILL = "FFCFF7E0"
GRIDL = Side(style="thin", color="FFDCE2EA")
GROUPL = Side(style="medium", color="FF8DA0B6")
VDIV = Side(style="medium", color="FF8DA0B6")
def money(c): c.number_format = '"$"#,##0.00'
def intfmt(c): c.number_format = '#,##0'
H_FILL, H_FONT = PatternFill("solid", fgColor=INK), Font(bold=True, color="FFFFFFFF", size=10)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
def med(xs):
    xs = [x for x in xs if x is not None]
    return statistics.median(xs) if xs else None

order = {"foodservice": 0, "appliance": 1, "beauty": 2}
prods_sorted = sorted(prods, key=lambda p: (order.get(p["line"], 9), p.get("group_name") or "", p["name"]))
wb = Workbook()

# ═══ 1. Overview ═══════════════════════════════════════════════════
ov = wb.active; ov.title = "Overview"; ov.sheet_view.showGridLines = False
ov.column_dimensions["A"].width = 2; ov.column_dimensions["B"].width = 27; ov.column_dimensions["C"].width = 74
ov["B2"] = "THE PORTAL"; ov["B2"].font = Font(bold=True, size=22, color=INK)
ov["B3"] = f"Product & Competitor Backup   ·   {datetime.date.today().strftime('%B %-d, %Y')}"; ov["B3"].font = Font(size=11, color=MUT)
r = 5
def ov_line(label, val, bold=False, hue=None, small=False):
    global r
    a = ov.cell(row=r, column=2, value=label); a.font = Font(bold=not small, color=SLATE, size=10 if small else 11)
    b = ov.cell(row=r, column=3, value=val); b.font = Font(bold=bold, color=hue or ("FF64748B" if small else "FF0F172A"), size=10 if small else 11)
    b.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
fs = [p for p in prods if p["line"] == "foodservice"]; ap = [p for p in prods if p["line"] == "appliance"]; be = [p for p in prods if p["line"] == "beauty"]
ov_line("What this is", "An offline snapshot of the Portal for reviewing and pricing products if the site is unavailable. Each product carries its full deal math and sits beside its own Amazon competitors.")
r += 1
ov_line("Live site", SITE, hue=LINK)
ov_line("Products", str(len(prods)), bold=True)
ov_line("   Foodservice", f"{len(fs)}   ready to sell — has costs, deal math, and competitor data", hue=HUE['foodservice'])
ov_line("   Appliances", f"{len(ap)}   no cost yet — set a target sell price to generate one", hue=HUE['appliance'])
ov_line("   Beauty", f"{len(be)}", hue=HUE['beauty'])
ov_line("Competitors", f"{len(comps)} verified listings across {len(cby)} products — each links to Amazon", bold=True)
r += 1
ov_line("Tabs", "")
ov_line("   Products & Competitors", "The working view. Each product's deal math is pinned on the left; its competitors line up to the right. Net and Margin are live formulas — type a new Target Sell and they update.")
ov_line("   Catalog", "The full 278-product index with specs, cost, competitor count, and a link to each live page.")
r += 1
ov_line("The deal math (left block, per product)", "")
for t, d in [("Target Sell", "the price we'd list at — pre-filled at the market median, editable"),
             ("FBA / Unit", "Amazon's real fulfillment fee (median of this product's competitors)"),
             ("Sell Fees (34%)", "referral 15% + ads 15% + returns 4% of the sell price"),
             ("Net / Unit", "Target Sell − our cost − sell fees − FBA"),
             ("Margin", "Net ÷ Target Sell   ·   green ≥ 15%,  amber 8–15%,  red < 8%")]:
    ov_line("   " + t, d, small=True)
r += 1
ov_line("Competitor columns (right)", "")
for t, d in [("Price", "current Amazon price — lowest in each group is shaded green"),
             ("BSR", "sales rank — lower is stronger"),
             ("Sold/Mo", "estimated units sold last month"),
             ("Reviews", "total ratings — how entrenched the listing is"),
             ("Their FBA", "that listing's own fulfillment fee")]:
    ov_line("   " + t, d, small=True)

# ═══ 2. Products & Competitors ═════════════════════════════════════
rv = wb.create_sheet("Products & Competitors"); rv.sheet_view.showGridLines = False
cols = ["Product", "Our Cost", "Target Sell", "FBA Fee", "Sell Fees 34%", "Net / Unit", "Margin",
        "Competitor  (click to open)", "Price", "BSR", "Sold/Mo", "Reviews", "Their FBA"]
wds = [30, 10, 11, 9, 11, 10, 9, 42, 9, 9, 8, 8, 9]
right_num = {2, 3, 4, 5, 6, 7, 9, 10, 11, 12, 13}
for i, (t, w) in enumerate(zip(cols, wds), 1):
    cc = rv.cell(row=1, column=i, value=t); cc.fill = H_FILL; cc.font = H_FONT
    cc.alignment = Alignment(horizontal="right" if i in right_num else "left", vertical="center", wrap_text=True)
    rv.column_dimensions[get_column_letter(i)].width = w
rv.row_dimensions[1].height = 40
rv.freeze_panes = "H2"          # pin header + the economics block
rv.print_title_rows = "1:1"
rr = 2; gi = 0
for p in [p for p in prods_sorted if cby.get(p["id"])]:
    grp = sorted(cby[p["id"]], key=lambda c: -(c.get("est_monthly_sales") or 0))
    gi += 1; band = PatternFill("solid", fgColor=BAND) if gi % 2 == 0 else None
    start = rr
    prices = [float(c["price"]) for c in grp if c.get("price") is not None]
    lowp = min(prices) if prices else None
    m = med(prices)
    fba = med([float(c["fba_pick_pack_fee"]) for c in grp if c.get("fba_pick_pack_fee") is not None])
    cost = costs.get(p["external_ref"])
    for c in grp:                   # competitors → columns 8-13
        cvals = [c["title"], c.get("price"), c.get("bsr"), c.get("est_monthly_sales"), c.get("review_count"), c.get("fba_pick_pack_fee")]
        for j, v in enumerate(cvals):
            ci = 8 + j
            cc = rv.cell(row=rr, column=ci, value=(float(v) if ci in (9, 13) and v is not None else v))
            cc.font = Font(size=10, color="FF0F172A")
            cc.alignment = LEFT if ci == 8 else RIGHT
            cc.border = Border(bottom=GRIDL, top=(GROUPL if rr == start else None))
            if band: cc.fill = band
        link = rv.cell(row=rr, column=8)
        if c.get("retail_url"): link.hyperlink = c["retail_url"]; link.font = Font(color=LINK, underline="single", size=10)
        money(rv.cell(row=rr, column=9)); money(rv.cell(row=rr, column=13))
        for ci in (10, 11, 12): intfmt(rv.cell(row=rr, column=ci))
        if lowp is not None and c.get("price") is not None and float(c["price"]) == lowp:
            pc = rv.cell(row=rr, column=9); pc.fill = PatternFill("solid", fgColor=LOWFILL); pc.font = Font(size=10, bold=True, color="FF065F46")
        rv.row_dimensions[rr].height = 32; rr += 1
    end = rr - 1
    for ci in range(1, 8):          # merge the economics block once per product
        if end > start: rv.merge_cells(start_row=start, start_column=ci, end_row=end, end_column=ci)
    pc = rv.cell(row=start, column=1, value=p["name"]); pc.font = Font(size=11, bold=True, color=HUE.get(p["line"], INK)); pc.alignment = Alignment(vertical="center", wrap_text=True)
    if cost is not None:            # live deal math — Net & Margin recompute if Target Sell is edited
        Cs, Bs, Ds, Es, Fs = f"C{start}", f"B{start}", f"D{start}", f"E{start}", f"F{start}"
        rv.cell(row=start, column=2, value=cost)
        rv.cell(row=start, column=3, value=round(m, 2) if m else None)      # target sell = market median (editable)
        rv.cell(row=start, column=4, value=round(fba, 2) if fba else None)  # our FBA = competitor median
        rv.cell(row=start, column=5, value=f"=0.34*{Cs}")
        rv.cell(row=start, column=6, value=f"={Cs}-{Bs}-{Es}-{Ds}")
        rv.cell(row=start, column=7, value=f'=IF({Cs}>0,{Fs}/{Cs},"")')
        for ci in (2, 3, 4, 5, 6): money(rv.cell(row=start, column=ci))
        rv.cell(row=start, column=7).number_format = "0.0%"
        sell = m or 0; net = sell - (cost or 0) - OPEX * sell - (fba or 0); marg = net / sell if sell else 0
        hue = "FF047857" if marg >= 0.15 else ("FFB45309" if marg >= 0.08 else "FFB91C1C")
        rv.cell(row=start, column=2).font = Font(size=11, bold=True, color="FF0F172A")
        for ci in (3, 4, 5): rv.cell(row=start, column=ci).font = Font(size=10, color=SLATE)
        rv.cell(row=start, column=6).font = Font(size=11, bold=True, color=hue)
        rv.cell(row=start, column=7).font = Font(size=11, bold=True, color=hue)
    for ci in range(2, 8): rv.cell(row=start, column=ci).alignment = CENTER
    for ci in range(1, 8):          # re-apply band + separators across the merged block
        for rw in range(start, end + 1):
            cell = rv.cell(row=rw, column=ci)
            cell.border = Border(top=GROUPL if rw == start else None, bottom=GRIDL if rw == end else None, right=(VDIV if ci == 7 else None))
            if band: cell.fill = band

# ═══ 3. Catalog ════════════════════════════════════════════════════
cat = wb.create_sheet("Catalog"); cat.sheet_view.showGridLines = False
cols = ["Line", "Category", "Product", "Model", "Key Specs", "Our Cost", "Competitors", "Live Page"]
wds = [13, 22, 42, 16, 50, 11, 12, 11]
for i, (t, w) in enumerate(zip(cols, wds), 1):
    cc = cat.cell(row=1, column=i, value=t); cc.fill = H_FILL; cc.font = H_FONT
    cc.alignment = RIGHT if i in (6, 7) else Alignment(horizontal="left", vertical="center")
    cat.column_dimensions[get_column_letter(i)].width = w
cat.row_dimensions[1].height = 26; cat.freeze_panes = "A2"; cat.print_title_rows = "1:1"
rr = 2
for p in prods_sorted:
    specs = "; ".join(f'{s["label"]}: {s["value"]}' for s in (p.get("specs") or [])[:5])
    ncomp = len(cby.get(p["id"], []))
    vals = [p["line"].title(), p.get("group_name") or "—", p["name"], p.get("model") or "—", specs,
            costs.get(p["external_ref"]), ncomp if ncomp else None, None]
    for ci, v in enumerate(vals, 1):
        cc = cat.cell(row=rr, column=ci, value=v); cc.border = Border(bottom=GRIDL)
        cc.alignment = RIGHT if ci in (6, 7) else LEFT
        cc.font = Font(size=10, color="FF0F172A")
        if rr % 2 == 0: cc.fill = PatternFill("solid", fgColor=BAND)
    cat.cell(row=rr, column=1).font = Font(size=10, bold=True, color=HUE.get(p["line"], INK))
    money(cat.cell(row=rr, column=6))
    lp = cat.cell(row=rr, column=8, value="Open ↗"); lp.hyperlink = f"{SITE}/p/{p['external_ref'].split(':')[1]}"
    lp.font = Font(color=LINK, underline="single", size=10); lp.alignment = LEFT
    cat.row_dimensions[rr].height = 30; rr += 1
cat.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{rr-1}"

for ws, land in [(ov, False), (rv, True), (cat, True)]:
    ws.page_setup.orientation = "landscape" if land else "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins.left = ws.page_margins.right = 0.3; ws.page_margins.top = ws.page_margins.bottom = 0.4

out = os.path.expanduser(f"~/Desktop/Portal-Backup-{datetime.date.today().isoformat()}.xlsx")
wb.save(out)
print("saved:", out)
print(f"products {len(prods)} | competitors {len(comps)} across {len(cby)} products")
